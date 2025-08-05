import os
import win32com.client
import re
import time as time_module
import sqlite3
import json
import xlrd
from pyxlsb2 import open_workbook
from openpyxl import load_workbook
import datetime
from datetime import datetime, timedelta, time
import random
import shutil
import psutil
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
import calendar
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext

# --- CONFIGURAÇÃO E SEGURANÇA ---
PRO_URL = "https://pro.manserv.com.br/login"
PRO_USER = "wesley.luz@manserv.com.br"  #user
PRO_PASS = "028885" #password

DIRETORIO_ATUAL = os.path.dirname(os.path.abspath(__file__))
PASTA_DOWNLOAD = os.path.abspath(os.path.join(DIRETORIO_ATUAL, "downloads"))
ARQUIVO_DESTINO = os.path.join(DIRETORIO_ATUAL, "exportacao.xlsx")
ARQUIVO_LIMPEZA = os.path.join(DIRETORIO_ATUAL, "cronograma_lc.xlsx")

if not os.path.exists(PASTA_DOWNLOAD):
    os.makedirs(PASTA_DOWNLOAD)

def coletar_e_salvar_dados():
    print("Iniciando o robô de coleta de dados...")
    #limpar_pasta_downloads()

    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("start-maximized")
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36"
    )

    prefs = {
        "download.default_directory": PASTA_DOWNLOAD,
        "download.prompt_for_download": False,
        "directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=chrome_options)

    # Remover a detecção do webdriver
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
        """
    })

    wait = WebDriverWait(driver, 30)

    try:
        print(f"Acessando {PRO_URL}...")
        driver.get(PRO_URL)
        print("Realizando login...")
        wait.until(EC.presence_of_element_located((By.ID, 'client-email'))).send_keys(PRO_USER)
        wait.until(EC.presence_of_element_located((By.ID, 'client-password'))).send_keys(PRO_PASS)
        wait.until(EC.element_to_be_clickable((By.ID, 'client-submit'))).click()
        print("Login realizado com sucesso!")
        print("Navegando até o relatório...")
        wait = WebDriverWait(driver, 10)
        dropdown_container = wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, ".selectize-control.form-control.single.plugin-restore_on_backspace")
        ))
        dropdown_container.click()
        dropdown_opcoes = wait.until(EC.visibility_of_element_located(
            (By.CSS_SELECTOR, ".selectize-dropdown")
        ))
        opcao = dropdown_opcoes.find_element(By.XPATH, ".//div[contains(text(),'MF - STIHL SERVIÇOS')]")
        opcao.click()
        driver.get("https://pro.manserv.com.br/operational/checklist-results-history")
        hoje = datetime.today()
        primeiro_dia = hoje.replace(day=1).strftime("%d/%m/%Y")
        ultimo_dia = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1]).strftime("%d/%m/%Y")
        script = f'''
        document.getElementById("beginDate").value = "{primeiro_dia}";
        document.getElementById("beginDate").dispatchEvent(new Event('change'));

        document.getElementById("endDate").value = "{ultimo_dia}";
        document.getElementById("endDate").dispatchEvent(new Event('change'));
        '''
        driver.execute_script(script)
        wait = WebDriverWait(driver, 10)
        botao_filtrar = wait.until(EC.element_to_be_clickable((By.ID, "button-filter")))
        botao_filtrar.click()
        wait.until(EC.element_to_be_clickable((By.ID, "button-export-excel"))).click()

        print("Aguardando o download do arquivo...")
        tempo_espera = 0
        arquivo_baixado = None
        while tempo_espera < 600:  # Espera por no máximo 300 segundos
            if any(f.endswith('.crdownload') for f in os.listdir(PASTA_DOWNLOAD)):
                time_module.sleep(1)
                tempo_espera += 1
                continue
            lista_arquivos = [
                f for f in os.listdir(PASTA_DOWNLOAD)
                if f.endswith(('.xls', '.xlsx')) and not f.startswith('~$')
            ]
            if lista_arquivos:
                arquivo_baixado = os.path.join(PASTA_DOWNLOAD, lista_arquivos[0])
                print(f"Download concluído: {arquivo_baixado}")
                break
            time_module.sleep(1)
            tempo_espera += 1
        if not arquivo_baixado:
            raise Exception("O download do arquivo demorou muito ou falhou.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        novo_nome = os.path.join(PASTA_DOWNLOAD, f"download_{timestamp}.xls")
        os.rename(arquivo_baixado, novo_nome)
        arquivo_baixado = novo_nome
        print("Processando a planilha...")
        df_novo = pd.read_excel(arquivo_baixado, skiprows=1, engine='xlrd')
        df_novo = df_novo.iloc[:, :11]
        df_novo.columns = [
            'id_resposta', 'id_empresa', 'id_checklist', 'checklist', 'data_inicio', 'data_fim',
            'id_ativo', 'ativo', 'qr_code', 'usuario', 'data_registro'
        ]
        print(df_novo.head())
        with pd.ExcelWriter(ARQUIVO_DESTINO, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_novo.to_excel(writer, sheet_name='Worksheet', startrow=1, header=False, index=False)
        print(f"Arquivo {os.path.basename(ARQUIVO_DESTINO)} atualizado com {len(df_novo)} registros.")
        print(f"Limpando o arquivo baixado: {os.path.basename(arquivo_baixado)}")
        tentar_excluir_arquivo(arquivo_baixado)
        df_final = pd.read_excel(ARQUIVO_DESTINO, sheet_name='Worksheet', skiprows=1)
        df_final.columns = [
            'id_resposta', 'id_empresa', 'id_checklist', 'checklist', 'data_inicio', 'data_fim',
            'id_ativo', 'ativo', 'qr_code', 'usuario', 'data_registro'
        ]
        salvar_dados_no_banco(df_final)

        wb = load_workbook(ARQUIVO_LIMPEZA)
        ws = wb["MSPRO_DB"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
            for cell in row:
                cell.value = None
        for row_idx, row in enumerate(df_final.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx).value = value
        wb.save(ARQUIVO_LIMPEZA)
        print("Dados gravados com sucesso a partir da célula A2, mantendo o cabeçalho e fórmulas.")
        print(ARQUIVO_LIMPEZA)

        #destino = r"C:\\Users\\manserv\\MANSERV\\Manserv - UT Stihl - Documentos\\Planejamento\\Cronogramas\\cronograma_lc.xlsx"
        #shutil.copy2(ARQUIVO_LIMPEZA, destino)
        #print("Arquivo substituído com sucesso no OneDrive.")

    except Exception as e:
        print(f"\nOcorreu um erro durante a automação: {e}")
        driver.save_screenshot("erro_screenshot.png")
        print("Um screenshot do erro foi salvo como 'erro_screenshot.png'.")
        raise

    finally:
        print("Fechando o navegador.")
        driver.quit()
        limpar_pasta_downloads()

def converter_xls_para_xlsx(caminho_xls):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    caminho_absoluto = os.path.abspath(caminho_xls)
    pasta, nome_arquivo = os.path.split(caminho_absoluto)
    nome_sem_extensao = os.path.splitext(nome_arquivo)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_convertido = os.path.join(pasta, f"{nome_sem_extensao}_{timestamp}.xlsx")

    if os.path.exists(caminho_convertido):
        os.remove(caminho_convertido)

    wb = None
    try:
        wb = excel.Workbooks.Open(caminho_absoluto)
        wb.SaveAs(caminho_convertido, FileFormat=51)
    finally:
        if wb:
            wb.Close(SaveChanges=False)
            del wb
        excel.Quit()
        del excel

    return caminho_convertido

def salvar_dados_no_banco(df):
    conn = sqlite3.connect("cronograma-lc.db")
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS checklists (
            id_resposta TEXT PRIMARY KEY,
            id_empresa TEXT,
            id_checklist TEXT,
            checklist TEXT,
            data_inicio TEXT,
            data_fim TEXT,
            id_ativo TEXT,
            ativo TEXT,
            qr_code TEXT,
            usuario TEXT,
            data_registro TEXT
        )
    """)

    for _, row in df.iterrows():
        cursor.execute("""
            INSERT OR REPLACE INTO checklists (
                id_resposta, id_empresa, id_checklist, checklist, data_inicio, data_fim,
                id_ativo, ativo, qr_code, usuario, data_registro
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            str(row.get('id_resposta', '')),
            str(row.get('id_empresa', '')),
            str(row.get('id_checklist', '')),
            str(row.get('checklist', '')),
            str(row.get('data_inicio', '')),
            str(row.get('data_fim', '')),
            str(row.get('id_ativo', '')),
            str(row.get('ativo', '')),
            str(row.get('qr_code', '')),
            str(row.get('usuario', '')),
            str(row.get('data_registro', ''))
        ))

    conn.commit()
    conn.close()
    print("✅ Dados salvos no banco cronograma-lc.db com sucesso!")


def ajustar_data_turno(data_hora, turno):
    hora = data_hora.time()
    if turno == "T1" and hora >= time(22, 35):
        return (data_hora + pd.Timedelta(days=1)).date()
    if turno == "T3E" and hora <= time(1, 9):
        return (data_hora - pd.Timedelta(days=1)).date()
    return data_hora.date()

TURNOS_HORARIOS = {
    "T1": (time(22, 35), time(6, 0)),
    "T2": (time(6, 0), time(14, 20)),
    "T3": (time(14, 20), time(22, 35)),
    "T2E": (time(6, 0), time(15, 48)),
    "T3E": (time(15, 48), time(1, 9)),
}

def gerar_faltando_json():
    xlsx_path = ARQUIVO_LIMPEZA
    output_path = os.path.join(DIRETORIO_ATUAL, "../frontend/faltando.json")

    sheets = pd.read_excel(xlsx_path, sheet_name=None)
    cronograma = sheets.get("Cronograma")
    leituras = sheets.get("MSPRO_DB")

    if cronograma is None or leituras is None:
        print("Erro: As abas 'Cronograma' e 'MSPRO_DB' não foram encontradas.")
        return

    dias_semana = ["SEG", "TER", "QUA", "QUI", "SEX", "SÁB", "DOM"]
    hoje = datetime.now()
    dia_atual = dias_semana[hoje.weekday()]

    faltando_por_turno = {}

    for turno, (inicio, fim) in TURNOS_HORARIOS.items():
        ambientes_hoje = cronograma[
            (cronograma[dia_atual].astype(str).str.strip().str.upper() == "X") &
            (cronograma["Turnos"].apply(lambda t: turno.upper() in str(t).upper()))
        ].copy()

        leituras["Data/Hora de Início"] = pd.to_datetime(
            leituras["Data/Hora de Início"], errors="coerce", dayfirst=True
        )
        leituras["Data_Logica"] = leituras["Data/Hora de Início"].apply(
            lambda dt: ajustar_data_turno(dt, turno) if not pd.isna(dt) else None
        )
        hoje_data = hoje.date()

        df_turno = leituras[leituras["Data_Logica"] == hoje_data]

        if inicio < fim:
            df_turno = df_turno[
                (df_turno["Data/Hora de Início"].dt.time >= inicio) &
                (df_turno["Data/Hora de Início"].dt.time <= fim)
            ]
        else:
            df_turno = df_turno[
                (df_turno["Data/Hora de Início"].dt.time >= inicio) |
                (df_turno["Data/Hora de Início"].dt.time <= fim)
            ]

        qrcodes_lidos = set(df_turno["QRCode"].astype(str)) if not df_turno.empty else set()

        ambientes_hoje["Lido"] = ambientes_hoje.apply(
            lambda row: (
                str(row["Local Instalação"]) in qrcodes_lidos or
                str(row["Arvore Prisma4 / Pro"]) in qrcodes_lidos
            ),
            axis=1
        )

        faltantes = ambientes_hoje[~ambientes_hoje["Lido"]]
        faltantes = faltantes.fillna('')
        faltando_por_turno[turno] = faltantes[["Local Instalação", "Arvore Prisma4 / Pro", "Descrição", "Turnos"]].to_dict(orient="records")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(faltando_por_turno, f, ensure_ascii=False, indent=2)

    print(f"faltando.json salvo em {output_path}")

def limpar_pasta_downloads():
    if os.path.exists(PASTA_DOWNLOAD):
        for arquivo in os.listdir(PASTA_DOWNLOAD):
            caminho = os.path.join(PASTA_DOWNLOAD, arquivo)
            if os.path.isfile(caminho):
                try:
                    os.remove(caminho)
                except:
                    print("Arquivo em uso.")
        print("Pasta downloads limpa.")

def kill_excel_processes():
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] and 'EXCEL.EXE' in proc.info['name'].upper():
            try:
                proc.kill()
                print(f"🔴 Processo {proc.info['pid']} (EXCEL.EXE) finalizado à força.")
            except Exception as e:
                print(f"Erro ao matar processo EXCEL.EXE: {e}")

def tentar_excluir_arquivo(caminho, tentativas=5, espera=1):
    import time
    for _ in range(tentativas):
        try:
            os.remove(caminho)
            print(f"Arquivo removido: {caminho}")
            return True
        except PermissionError:
            print(f"Aguardando liberação do arquivo: {caminho}")
            time.sleep(espera)
    print(f"Não foi possível excluir: {caminho}")
    return False

if __name__ == '__main__':
    max_tentativas = 3
    for tentativa in range(1, max_tentativas + 1):
        print(f"\n🔄 Tentativa {tentativa} de {max_tentativas}...\n")
        try:
            coletar_e_salvar_dados()
            gerar_faltando_json()
            print("✅ Execução concluída com sucesso.")
            break
        except Exception as e:
            print(f"❌ Erro na tentativa {tentativa}: {e}")
            if tentativa < max_tentativas:
                print("⏳ Aguardando 30 segundos antes de tentar novamente...")
                time_module.sleep(30)
            else:
                print("🚫 Todas as tentativas falharam. Encerrando.")
        limpar_pasta_downloads()