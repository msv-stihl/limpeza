#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de teste para verificar o funcionamento do sistema
"""

import os
import sys
import json
import sqlite3
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

# Configurar logging para testes
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class SystemTester:
    """Classe para testar o sistema de limpeza"""
    
    def __init__(self):
        self.project_dir = Path(__file__).parent.absolute()
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.test_count = 0
    
    def log_test(self, test_name, success, message=""):
        """Registra resultado de um teste"""
        self.test_count += 1
        
        if success:
            self.success_count += 1
            logger.info(f"✅ {test_name}: PASSOU {message}")
        else:
            self.errors.append(f"{test_name}: {message}")
            logger.error(f"❌ {test_name}: FALHOU {message}")
    
    def log_warning(self, test_name, message):
        """Registra um aviso"""
        self.warnings.append(f"{test_name}: {message}")
        logger.warning(f"⚠️  {test_name}: {message}")
    
    def test_python_environment(self):
        """Testa o ambiente Python"""
        logger.info("=== Testando Ambiente Python ===")
        
        # Versão do Python
        python_version = sys.version_info
        self.log_test(
            "Versão Python",
            python_version >= (3, 6),
            f"v{python_version.major}.{python_version.minor}.{python_version.micro}"
        )
        
        # Módulos necessários
        required_modules = [
            'requests', 'pandas', 'openpyxl', 'xlrd', 'beautifulsoup4',
            'schedule', 'git', 'python-dotenv', 'psutil'
        ]
        
        for module in required_modules:
            try:
                if module == 'beautifulsoup4':
                    import bs4
                elif module == 'python-dotenv':
                    import dotenv
                else:
                    __import__(module.replace('-', '_'))
                self.log_test(f"Módulo {module}", True, "importado com sucesso")
            except ImportError as e:
                self.log_test(f"Módulo {module}", False, f"não encontrado: {e}")
    
    def test_file_structure(self):
        """Testa a estrutura de arquivos"""
        logger.info("=== Testando Estrutura de Arquivos ===")
        
        required_files = [
            'main.py',
            'coletor_linux.py',
            'git_manager.py',
            'requirements.txt',
            '.env.example',
            'install_linux.sh',
            'setup_cron.sh',
            'README.md',
            '.gitignore'
        ]
        
        for file_name in required_files:
            file_path = self.project_dir / file_name
            self.log_test(
                f"Arquivo {file_name}",
                file_path.exists(),
                f"encontrado em {file_path}" if file_path.exists() else "não encontrado"
            )
        
        # Diretórios
        required_dirs = ['frontend', 'downloads']
        
        for dir_name in required_dirs:
            dir_path = self.project_dir / dir_name
            if not dir_path.exists():
                dir_path.mkdir(exist_ok=True)
                self.log_warning(f"Diretório {dir_name}", "criado automaticamente")
            else:
                self.log_test(f"Diretório {dir_name}", True, "encontrado")
        
        # Arquivos do frontend
        frontend_files = ['index.html', 'faltando.html', 'main.js', 'styles.css']
        
        for file_name in frontend_files:
            file_path = self.project_dir / 'frontend' / file_name
            self.log_test(
                f"Frontend {file_name}",
                file_path.exists(),
                "encontrado" if file_path.exists() else "não encontrado"
            )
    
    def test_configuration(self):
        """Testa configurações"""
        logger.info("=== Testando Configurações ===")
        
        # Arquivo .env
        env_file = self.project_dir / '.env'
        env_example = self.project_dir / '.env.example'
        
        if env_file.exists():
            self.log_test("Arquivo .env", True, "encontrado")
            
            # Verificar variáveis essenciais
            try:
                from dotenv import load_dotenv
                load_dotenv(env_file)
                
                pro_user = os.getenv('PRO_USER')
                pro_pass = os.getenv('PRO_PASS')
                
                self.log_test(
                    "PRO_USER configurado",
                    bool(pro_user and pro_user != 'seu_usuario@manserv.com.br'),
                    f"valor: {pro_user[:10]}..." if pro_user else "não configurado"
                )
                
                self.log_test(
                    "PRO_PASS configurado",
                    bool(pro_pass and pro_pass != 'sua_senha'),
                    "configurado" if pro_pass else "não configurado"
                )
                
            except Exception as e:
                self.log_test("Leitura .env", False, str(e))
        
        elif env_example.exists():
            self.log_warning("Arquivo .env", "não encontrado, mas .env.example existe")
        else:
            self.log_test("Arquivo .env", False, "não encontrado")
    
    def test_data_files(self):
        """Testa arquivos de dados"""
        logger.info("=== Testando Arquivos de Dados ===")
        
        # Arquivos Excel
        excel_files = ['exportacao.xlsx', 'cronograma_lc.xlsx']
        
        for file_name in excel_files:
            file_path = self.project_dir / file_name
            
            if file_path.exists():
                try:
                    # Tentar ler o arquivo
                    wb = load_workbook(file_path, read_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                    
                    self.log_test(
                        f"Excel {file_name}",
                        True,
                        f"válido com {len(sheets)} abas: {', '.join(sheets[:3])}"
                    )
                    
                except Exception as e:
                    self.log_test(f"Excel {file_name}", False, f"erro ao ler: {e}")
            else:
                self.log_warning(f"Excel {file_name}", "não encontrado (será criado na execução)")
        
        # Banco de dados
        db_file = self.project_dir / 'cronograma-lc.db'
        
        if db_file.exists():
            try:
                conn = sqlite3.connect(db_file)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                tables = cursor.fetchall()
                conn.close()
                
                self.log_test(
                    "Banco SQLite",
                    True,
                    f"válido com {len(tables)} tabelas"
                )
                
            except Exception as e:
                self.log_test("Banco SQLite", False, f"erro ao ler: {e}")
        else:
            self.log_warning("Banco SQLite", "não encontrado (será criado na execução)")
        
        # JSON do frontend
        json_file = self.project_dir / 'frontend' / 'faltando.json'
        
        if json_file.exists():
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                self.log_test(
                    "JSON faltando",
                    True,
                    f"válido com {len(data)} turnos"
                )
                
            except Exception as e:
                self.log_test("JSON faltando", False, f"erro ao ler: {e}")
        else:
            self.log_warning("JSON faltando", "não encontrado (será criado na execução)")
    
    def test_scripts_permissions(self):
        """Testa permissões dos scripts (apenas Linux)"""
        if os.name != 'posix':
            self.log_warning("Permissões scripts", "teste pulado (não é Linux)")
            return
        
        logger.info("=== Testando Permissões dos Scripts ===")
        
        scripts = ['install_linux.sh', 'setup_cron.sh']
        
        for script in scripts:
            script_path = self.project_dir / script
            
            if script_path.exists():
                is_executable = os.access(script_path, os.X_OK)
                self.log_test(
                    f"Script {script} executável",
                    is_executable,
                    "sim" if is_executable else "não (execute: chmod +x)"
                )
            else:
                self.log_test(f"Script {script}", False, "não encontrado")
    
    def test_network_connectivity(self):
        """Testa conectividade de rede"""
        logger.info("=== Testando Conectividade ===")
        
        try:
            import requests
            
            # Testar acesso ao Pro Manserv
            try:
                response = requests.get('https://pro.manserv.com.br', timeout=10)
                self.log_test(
                    "Acesso Pro Manserv",
                    response.status_code == 200,
                    f"status {response.status_code}"
                )
            except Exception as e:
                self.log_test("Acesso Pro Manserv", False, str(e))
            
            # Testar acesso ao GitHub
            try:
                response = requests.get('https://github.com', timeout=10)
                self.log_test(
                    "Acesso GitHub",
                    response.status_code == 200,
                    f"status {response.status_code}"
                )
            except Exception as e:
                self.log_test("Acesso GitHub", False, str(e))
                
        except ImportError:
            self.log_test("Teste conectividade", False, "módulo requests não disponível")
    
    def test_main_script(self):
        """Testa o script principal"""
        logger.info("=== Testando Script Principal ===")
        
        try:
            # Importar módulos principais
            sys.path.append(str(self.project_dir))
            
            try:
                import main
                self.log_test("Import main.py", True, "importado com sucesso")
            except Exception as e:
                self.log_test("Import main.py", False, str(e))
            
            try:
                import coletor_linux
                self.log_test("Import coletor_linux.py", True, "importado com sucesso")
            except Exception as e:
                self.log_test("Import coletor_linux.py", False, str(e))
            
            try:
                import git_manager
                self.log_test("Import git_manager.py", True, "importado com sucesso")
            except Exception as e:
                self.log_test("Import git_manager.py", False, str(e))
                
        except Exception as e:
            self.log_test("Teste scripts", False, f"erro geral: {e}")
    
    def run_all_tests(self):
        """Executa todos os testes"""
        logger.info("🚀 Iniciando testes do sistema...")
        logger.info(f"Diretório do projeto: {self.project_dir}")
        
        # Executar todos os testes
        self.test_python_environment()
        self.test_file_structure()
        self.test_configuration()
        self.test_data_files()
        self.test_scripts_permissions()
        self.test_network_connectivity()
        self.test_main_script()
        
        # Relatório final
        logger.info("\n" + "="*60)
        logger.info("📊 RELATÓRIO FINAL DOS TESTES")
        logger.info("="*60)
        
        success_rate = (self.success_count / self.test_count * 100) if self.test_count > 0 else 0
        
        logger.info(f"✅ Testes passaram: {self.success_count}/{self.test_count} ({success_rate:.1f}%)")
        
        if self.warnings:
            logger.info(f"⚠️  Avisos: {len(self.warnings)}")
            for warning in self.warnings:
                logger.warning(f"   - {warning}")
        
        if self.errors:
            logger.info(f"❌ Erros: {len(self.errors)}")
            for error in self.errors:
                logger.error(f"   - {error}")
        
        # Recomendações
        logger.info("\n📋 PRÓXIMOS PASSOS:")
        
        if not (self.project_dir / '.env').exists():
            logger.info("1. Crie o arquivo .env com suas credenciais")
        
        if self.errors:
            logger.info("2. Corrija os erros listados acima")
        
        if success_rate >= 80:
            logger.info("3. Execute: python3 main.py --action status")
            logger.info("4. Configure o cron: ./setup_cron.sh")
        else:
            logger.info("3. Execute novamente após corrigir os problemas")
        
        logger.info("\n🎯 Sistema pronto para uso!" if success_rate >= 80 else "\n🔧 Sistema precisa de ajustes")
        
        return success_rate >= 80

def main():
    """Função principal"""
    tester = SystemTester()
    success = tester.run_all_tests()
    
    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()