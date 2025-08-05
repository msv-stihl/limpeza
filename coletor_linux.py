#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Coletor de dados para sistema de limpeza - Versão Linux
Compatível com cron e sem dependências do Windows
"""

import os
import re
import time
import sqlite3
import json
import requests
from datetime import datetime, timedelta, time as time_obj
import calendar
import shutil
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import logging
from pathlib import Path
import tempfile
from dotenv import load_dotenv

# Carregar variáveis de ambiente
load_dotenv()

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('coletor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# --- CONFIGURAÇÃO E SEGURANÇA ---
PRO_URL = "https://pro.manserv.com.br"
PRO_LOGIN_URL = f"{PRO_URL}/login"
PRO_CHECKLIST_URL = f"{PRO_URL}/operational/checklist-results-history"
PRO_EXPORT_URL = f"{PRO_URL}/operational/checklist-results-history/export"

# Credenciais (usar variáveis de ambiente em produção)
PRO_USER = os.getenv('PRO_USER', 'wesley.luz@manserv.com.br')
PRO_PASS = os.getenv('PRO_PASS', '028885')

# Diretórios
DIRETORIO_ATUAL = Path(__file__).parent.absolute()
PASTA_DOWNLOAD = DIRETORIO_ATUAL / "downloads"
ARQUIVO_DESTINO = DIRETORIO_ATUAL / "exportacao.xlsx"
ARQUIVO_LIMPEZA = DIRETORIO_ATUAL / "cronograma_lc.xlsx"
ARQUIVO_JSON = DIRETORIO_ATUAL / "frontend" / "faltando.json"

# Criar diretórios se não existirem
PASTA_DOWNLOAD.mkdir(exist_ok=True)
(DIRETORIO_ATUAL / "frontend").mkdir(exist_ok=True)

class ProManservCollector:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })
        
    def login(self):
        """Realiza login no sistema Pro Manserv"""
        try:
            logger.info("Acessando página de login...")
            
            # Primeiro, acessar a página de login para obter tokens CSRF se necessário
            response = self.session.get(PRO_LOGIN_URL)
            response.raise_for_status()
            
            # Parse da página para extrair tokens se necessário
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Dados do login
            login_data = {
                'client-email': PRO_USER,
                'client-password': PRO_PASS
            }
            
            # Verificar se há token CSRF
            csrf_token = soup.find('input', {'name': '_token'})
            if csrf_token:
                login_data['_token'] = csrf_token.get('value')
            
            logger.info("Realizando login...")
            response = self.session.post(PRO_LOGIN_URL, data=login_data)
            
            # Verificar se o login foi bem-sucedido
            if 'dashboard' in response.url.lower() or response.status_code == 200:
                logger.info("Login realizado com sucesso!")
                return True
            else:
                logger.error(f"Falha no login. Status: {response.status_code}")
                return False
                
        except Exception as e:
            logger.error(f"Erro durante o login: {e}")
            return False
    
    def get_checklist_data(self):
        """Obtém dados do checklist via requisição HTTP"""
        try:
            logger.info("Acessando página de relatórios...")
            
            # Acessar a página de relatórios
            response = self.session.get(PRO_CHECKLIST_URL)
            response.raise_for_status()
            
            # Calcular datas do mês atual
            hoje = datetime.today()
            primeiro_dia = hoje.replace(day=1).strftime("%Y-%m-%d")
            ultimo_dia = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1]).strftime("%Y-%m-%d")
            
            logger.info(f"Buscando dados de {primeiro_dia} até {ultimo_dia}")
            
            # Parse da página para obter parâmetros necessários
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Dados para o filtro
            filter_data = {
                'beginDate': primeiro_dia,
                'endDate': ultimo_dia,
                'company': 'MF - STIHL SERVIÇOS'  # Ajustar conforme necessário
            }
            
            # Verificar se há token CSRF
            csrf_token = soup.find('input', {'name': '_token'})
            if csrf_token:
                filter_data['_token'] = csrf_token.get('value')
            
            # Aplicar filtro
            logger.info("Aplicando filtros...")
            filter_response = self.session.post(PRO_CHECKLIST_URL, data=filter_data)
            filter_response.raise_for_status()
            
            # Tentar fazer download direto do Excel
            logger.info("Fazendo download do relatório...")
            
            # Dados para exportação
            export_data = filter_data.copy()
            export_data['format'] = 'excel'
            
            download_response = self.session.post(PRO_EXPORT_URL, data=export_data)
            
            if download_response.status_code == 200:
                # Salvar arquivo temporário
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_file = PASTA_DOWNLOAD / f"checklist_{timestamp}.xls"
                
                with open(temp_file, 'wb') as f:
                    f.write(download_response.content)
                
                logger.info(f"Arquivo baixado: {temp_file}")
                return temp_file
            else:
                logger.error(f"Erro no download: {download_response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"Erro ao obter dados do checklist: {e}")
            return None
    
    def process_excel_file(self, file_path):
        """Processa o arquivo Excel baixado"""
        try:
            logger.info("Processando arquivo Excel...")
            
            # Ler o arquivo Excel
            df_novo = pd.read_excel(file_path, skiprows=1, engine='xlrd')
            
            # Selecionar apenas as primeiras 11 colunas
            df_novo = df_novo.iloc[:, :11]
            
            # Definir nomes das colunas
            df_novo.columns = [
                'id_resposta', 'id_empresa', 'id_checklist', 'checklist', 'data_inicio', 'data_fim',
                'id_ativo', 'ativo', 'qr_code', 'usuario', 'data_registro'
            ]
            
            logger.info(f"Processados {len(df_novo)} registros")
            
            # Atualizar arquivo de exportação
            self.update_export_file(df_novo)
            
            # Salvar no banco de dados
            self.save_to_database(df_novo)
            
            # Atualizar cronograma
            self.update_cronograma(df_novo)
            
            # Limpar arquivo temporário
            if file_path.exists():
                file_path.unlink()
                logger.info(f"Arquivo temporário removido: {file_path}")
            
            return df_novo
            
        except Exception as e:
            logger.error(f"Erro ao processar arquivo Excel: {e}")
            return None
    
    def update_export_file(self, df_novo):
        """Atualiza o arquivo de exportação"""
        try:
            with pd.ExcelWriter(ARQUIVO_DESTINO, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_novo.to_excel(writer, sheet_name='Worksheet', startrow=1, header=False, index=False)
            
            logger.info(f"Arquivo {ARQUIVO_DESTINO.name} atualizado com {len(df_novo)} registros")
            
        except Exception as e:
            logger.error(f"Erro ao atualizar arquivo de exportação: {e}")
    
    def save_to_database(self, df):
        """Salva dados no banco SQLite"""
        try:
            conn = sqlite3.connect("cronograma-lc.db")
            cursor = conn.cursor()
            
            # Criar tabela se não existir
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS checklists (
                    id_resposta TEXT PRIMARY KEY,
                    id_empresa TEXT,
                    id_checklist TEXT,
                    checklist TEXT,
                    data_inicio TEXT,
                    data_fim TEXT,
                    id_ativo TEXT,
                    ativo TEXT,
                    qr_code TEXT,
                    usuario TEXT,
                    data_registro TEXT
                )
            """)
            
            # Inserir dados
            for _, row in df.iterrows():
                cursor.execute("""
                    INSERT OR REPLACE INTO checklists (
                        id_resposta, id_empresa, id_checklist, checklist, data_inicio, data_fim,
                        id_ativo, ativo, qr_code, usuario, data_registro
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    str(row.get('id_resposta', '')),
                    str(row.get('id_empresa', '')),
                    str(row.get('id_checklist', '')),
                    str(row.get('checklist', '')),
                    str(row.get('data_inicio', '')),
                    str(row.get('data_fim', '')),
                    str(row.get('id_ativo', '')),
                    str(row.get('ativo', '')),
                    str(row.get('qr_code', '')),
                    str(row.get('usuario', '')),
                    str(row.get('data_registro', ''))
                ))
            
            conn.commit()
            conn.close()
            logger.info("Dados salvos no banco cronograma-lc.db com sucesso!")
            
        except Exception as e:
            logger.error(f"Erro ao salvar no banco de dados: {e}")
    
    def update_cronograma(self, df_final):
        """Atualiza o arquivo de cronograma"""
        try:
            if not ARQUIVO_LIMPEZA.exists():
                logger.warning(f"Arquivo {ARQUIVO_LIMPEZA} não encontrado")
                return
            
            wb = load_workbook(ARQUIVO_LIMPEZA)
            
            if "MSPRO_DB" not in wb.sheetnames:
                logger.warning("Aba MSPRO_DB não encontrada no cronograma")
                return
            
            ws = wb["MSPRO_DB"]
            
            # Limpar dados existentes (mantendo cabeçalho)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=11):
                for cell in row:
                    cell.value = None
            
            # Inserir novos dados
            for row_idx, row in enumerate(df_final.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx).value = value
            
            wb.save(ARQUIVO_LIMPEZA)
            logger.info("Cronograma atualizado com sucesso")
            
        except Exception as e:
            logger.error(f"Erro ao atualizar cronograma: {e}")
    
    def generate_faltando_json(self):
        """Gera o arquivo JSON com ambientes faltantes"""
        try:
            if not ARQUIVO_LIMPEZA.exists():
                logger.warning(f"Arquivo {ARQUIVO_LIMPEZA} não encontrado")
                return
            
            logger.info("Gerando arquivo faltando.json...")
            
            # Ler planilhas
            sheets = pd.read_excel(ARQUIVO_LIMPEZA, sheet_name=None)
            cronograma = sheets.get("Cronograma")
            leituras = sheets.get("MSPRO_DB")
            
            if cronograma is None or leituras is None:
                logger.error("Abas 'Cronograma' e 'MSPRO_DB' não encontradas")
                return
            
            # Definir turnos e horários
            TURNOS_HORARIOS = {
                "T1": (time_obj(22, 35), time_obj(6, 0)),
                "T2": (time_obj(6, 0), time_obj(14, 20)),
                "T3": (time_obj(14, 20), time_obj(22, 35)),
                "T2E": (time_obj(6, 0), time_obj(15, 48)),
                "T3E": (time_obj(15, 48), time_obj(1, 9)),
            }
            
            dias_semana = ["SEG", "TER", "QUA", "QUI", "SEX", "SÁB", "DOM"]
            hoje = datetime.now()
            dia_atual = dias_semana[hoje.weekday()]
            
            faltando_por_turno = {}
            
            for turno, (inicio, fim) in TURNOS_HORARIOS.items():
                # Ambientes programados para hoje neste turno
                ambientes_hoje = cronograma[
                    (cronograma[dia_atual].astype(str).str.strip().str.upper() == "X") &
                    (cronograma["Turnos"].apply(lambda t: turno.upper() in str(t).upper()))
                ].copy()
                
                # Processar leituras
                leituras["Data/Hora de Início"] = pd.to_datetime(
                    leituras["Data/Hora de Início"], errors="coerce", dayfirst=True
                )
                
                leituras["Data_Logica"] = leituras["Data/Hora de Início"].apply(
                    lambda dt: self.ajustar_data_turno(dt, turno) if not pd.isna(dt) else None
                )
                
                hoje_data = hoje.date()
                df_turno = leituras[leituras["Data_Logica"] == hoje_data]
                
                # Filtrar por horário do turno
                if inicio < fim:
                    df_turno = df_turno[
                        (df_turno["Data/Hora de Início"].dt.time >= inicio) &
                        (df_turno["Data/Hora de Início"].dt.time <= fim)
                    ]
                else:
                    df_turno = df_turno[
                        (df_turno["Data/Hora de Início"].dt.time >= inicio) |
                        (df_turno["Data/Hora de Início"].dt.time <= fim)
                    ]
                
                # QR codes lidos
                qrcodes_lidos = set(df_turno["QRCode"].astype(str)) if not df_turno.empty else set()
                
                # Verificar quais ambientes foram lidos
                ambientes_hoje["Lido"] = ambientes_hoje.apply(
                    lambda row: (
                        str(row["Local Instalação"]) in qrcodes_lidos or
                        str(row["Arvore Prisma4 / Pro"]) in qrcodes_lidos
                    ),
                    axis=1
                )
                
                # Ambientes faltantes
                faltantes = ambientes_hoje[~ambientes_hoje["Lido"]]
                faltantes = faltantes.fillna('')
                
                faltando_por_turno[turno] = faltantes[
                    ["Local Instalação", "Arvore Prisma4 / Pro", "Descrição", "Turnos"]
                ].to_dict(orient="records")
            
            # Salvar JSON
            with open(ARQUIVO_JSON, "w", encoding="utf-8") as f:
                json.dump(faltando_por_turno, f, ensure_ascii=False, indent=2)
            
            logger.info(f"Arquivo faltando.json salvo em {ARQUIVO_JSON}")
            
        except Exception as e:
            logger.error(f"Erro ao gerar faltando.json: {e}")
    
    def ajustar_data_turno(self, data_hora, turno):
        """Ajusta a data baseada no turno"""
        if pd.isna(data_hora):
            return None
            
        hora = data_hora.time()
        
        if turno == "T1" and hora >= time_obj(22, 35):
            return (data_hora + pd.Timedelta(days=1)).date()
        if turno == "T3E" and hora <= time_obj(1, 9):
            return (data_hora - pd.Timedelta(days=1)).date()
        
        return data_hora.date()
    
    def cleanup_downloads(self):
        """Limpa a pasta de downloads"""
        try:
            if PASTA_DOWNLOAD.exists():
                for arquivo in PASTA_DOWNLOAD.iterdir():
                    if arquivo.is_file():
                        arquivo.unlink()
                logger.info("Pasta downloads limpa")
        except Exception as e:
            logger.error(f"Erro ao limpar downloads: {e}")
    
    def run(self):
        """Executa o processo completo de coleta"""
        logger.info("Iniciando coleta de dados...")
        
        try:
            # Login
            if not self.login():
                raise Exception("Falha no login")
            
            # Obter dados
            file_path = self.get_checklist_data()
            if not file_path:
                raise Exception("Falha ao obter dados")
            
            # Processar arquivo
            df_result = self.process_excel_file(file_path)
            if df_result is None:
                raise Exception("Falha ao processar arquivo")
            
            # Gerar JSON
            self.generate_faltando_json()
            
            logger.info("Coleta concluída com sucesso!")
            return True
            
        except Exception as e:
            logger.error(f"Erro durante a coleta: {e}")
            return False
        
        finally:
            self.cleanup_downloads()

def main():
    """Função principal"""
    max_tentativas = 3
    
    for tentativa in range(1, max_tentativas + 1):
        logger.info(f"Tentativa {tentativa} de {max_tentativas}")
        
        collector = ProManservCollector()
        
        if collector.run():
            logger.info("Execução concluída com sucesso")
            break
        else:
            if tentativa < max_tentativas:
                logger.info("Aguardando 30 segundos antes de tentar novamente...")
                time.sleep(30)
            else:
                logger.error("Todas as tentativas falharam")
                return False
    
    return True

if __name__ == '__main__':
    main()